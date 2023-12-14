import os
from views import *
from datetime import *
from const import *
import shutil

import openpyxl
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def create_Facture_xlsx(data: list, output_file: str, title, numero_vente: int) -> None:
    # Création du style de la police
    header_font = Font(bold=True, color="FFFFFF")  # Couleur blanche
    # Création du style de remplissage pour la couleur de fond
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Couleur rouge
    # Création du style de bordure pour l'en-tête
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value:  # Vérifier si la cellule n'est pas vide
                cell.border = thin_border  # Appliquer le style de bordure

    for cell in ws[1]:  # Parcourir chaque cellule de la première ligne
        cell.font = header_font  # Appliquer le style de police à chaque cellule
        cell.fill = header_fill  # Appliquer le style de remplissage à chaque cellule

    wb.save(output_file)  # Enregistrer le classeur Excel

    convert_to_pdf(output_file, f"Fact_{numero_vente}.pdf", title)  # Appeler la fonction de conversion vers PDF

def convert_to_pdf(input_file: str, output_file: str, title: str) -> None:
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    pdf = FPDF(orientation='L')
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Titre de la facture
    pdf.set_font("Arial", 'B', 16)
    pdf.multi_cell(0, 10, title, 0, 'C')
    pdf.ln(10)

    # Tableau avec les données
    row_height = 10
    col_width = 40
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            pdf.set_fill_color(0, 0, 0)  # Couleur de fond pour la première ligne
            pdf.set_text_color(255, 255, 255)  # Couleur du texte en blanc pour la première ligne
            pdf.set_x((pdf.w - (col_width * len(row))) / 2)  # Centrer le tableau
            for item in row:
                pdf.cell(col_width, row_height, str(item), 1, 0, 'C', 1)  # Centrage et couleur de fond pour la première ligne
            pdf.ln(row_height)
            first_row = False
        else:
            pdf.set_text_color(0, 0, 0)  # Réinitialiser la couleur du texte à noir pour les lignes suivantes
            pdf.set_x((pdf.w - (col_width * len(row))) / 2)  # Centrer le tableau
            for i, item in enumerate(row):
                if i == 5:  # Vérifier si c'est la sixième colonne (index 5)
                    pdf.multi_cell(50, row_height, str(item), 1)  # Utiliser multi_cell uniquement pour la sixième colonne
                else:
                    pdf.cell(col_width, row_height, str(item), 1, 0, 'C')  # Centrage pour les autres colonnes
            pdf.ln(row_height)

    pdf.output(output_file)

def get_file_content(file_path: str) -> list:
    list_products = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for product in f.read().splitlines():
            list_products.append(product.split(','))
        return list_products


def get_available_products(list_products: list) -> list:
    available_products = []
    available_products.insert(0, list_products[0])
    for product in list_products[1:]:
        if int(product[3]) > 0:
            available_products.append(product)
    return available_products


def get_product_by_id(list_product: list, IdP: int) -> list:
    for product in list_product[1:]:
        if IdP == int(product[0]):
            return product
    return None

def get_all_id_from_available(list_available: list) ->list:
    id_list = []
    for p in list_available[1:]:
        id_list.append(str(p[0]))
    return id_list

def get_id(list_product: list) -> int:
    id_put = ""
    while not (id_put.isdigit() and (id_put in get_all_id_from_available(list_product) or id_put == "0")):
        display_product(list_product)
        id_put = input("Entrez l'ID du produit : ")
        os.system("cls")
    return int(id_put)

def get_quantity(product_choosen: list) -> int:
    while True:
        print(f"Produit choisi : {product_choosen[1]}")
        qte = input("Quantité ? -> ")
        if not qte.isdigit():
            print("Veuillez entrer un nombre entier.")
            continue
        qte = int(qte)
        stock_quantity = int(product_choosen[3])

        if qte > stock_quantity:
            print(f"La quantité ne doit pas dépasser {stock_quantity}.")
        else:
            return qte

def ajouter_au_panier(panier, list_product, id_produit, quantite):
    product = get_product_by_id(list_product, id_produit)
    total = int(quantite) * int(product[2])
    description = product[4]
    panier.append([id_produit, product[1], quantite, product[2],description, total])
    print("Produit ajouté au panier.")

def supprimer_du_panier(panier, id_produit):
    for produit in panier[1:len(panier)-1]:
        if int(produit[0]) == id_produit:
            panier.remove(produit)
            print("Produit supprimé du panier.")
            return
    print("L'ID du produit n'a pas été trouvé dans le panier.")

def vider_panier(panier: list) -> list:
    panier = panier[:1]  # L'en-tête du panier
    return panier

def modifier_commande(panier: list, list_product: list, rep: int, somme):
    match(rep):
        case 1:
            main_loop(panier, list_product, somme)
            # id = get_id(list_product)
            # qte = get_quantity(get_product_by_id(list_product, id))
            # ajouter_au_panier(panier, list_product, id, qte)
        case 2:
            print("Voici votre panier actuel : ")
            id_produit = get_id(panier)
            print(type(panier[0][0]))
            supprimer_du_panier(panier, id_produit)
        case 3:
            panier = vider_panier(panier)

# def annuler_commande(panier):
#     print("Voici votre panier actuel : ")
#     display_product(panier)  
#     choix = input("Voulez-vous annuler une commande ? [oui/non] ")
#     if choix.lower() == "oui":
#     # vider_panier(panier)
#         id_produit = input("Entrez l'ID du produit à annuler : ")
#     for produit in panier:
#         if str(produit[0]) == id_produit:
#             panier.remove(produit)
#             print("La commande a été annulée.")
#             return
#     print("L'ID du produit n'a pas été trouvé dans le panier.")

def annuler_commande(panier) -> None:
    panier = vider_panier(panier)

def get_answer():
    while True:
        reponse = input("->")
        if not reponse.isdigit():
            continue
        else:
            return int(reponse) 


def get_answer_2(min: int, max :int):
    while True:
        reponse = get_answer()
        if reponse < min or reponse > max:
            continue
        else:
            return reponse
    


def update_product_database(product_id, quantity_sold, all_products_file, file_name: str):
    # Mettre à jour la quantité en stock du produit dans la liste de tous les produits
    for product in all_products_file:
        if product[0] == str(product_id):
            product[3] = str(int(product[3]) - quantity_sold)
            break

    # Maintenant, écrire la liste mise à jour dans le fichier de tous les produits
    with open(file_name, "w", encoding="utf-8") as file:
        for product in all_products_file:
            file.writelines(",".join(product))
            file.write("\n")
def get_position(panier, product):
    for i in range(len(panier)):
        if str(panier[i][0]) == product[0]:
            return i
    return None



def main_loop(panier, available_products: list, somme: int = 0):
    while True:
        os.system("cls")
        print("-"*5, "BIENVENUE DANS NOTRE BOUTIQUE", "-"*5)
        id = get_id(available_products)
        product = get_product_by_id(available_products, id)
        if not product == None:
            if not (get_position(panier, product) == None):
                pos = get_position(panier, product)
                qte = get_quantity(product)
                panier[pos][2] += qte
                panier[pos][5] += qte * int(product[2])
            else:
                qte = get_quantity(product)
                if qte == 0:
                    continue
                total = int(qte) * int(product[2])
                panier.append([id, product[1], qte, product[2], product[4],total])
            total = int(qte) * int(product[2])
            somme += total
            while True:
                rep = input("Voulez-vous ajouter un autre produit? [oui/non] ")
                if rep.upper() in ["OUI", "NON"]:
                    break
            product[3]= str(int(product[3]) - qte)
            if rep.upper() == "NON":
                return panier
                
            

def input_choice(data: str) -> str:
    while True:
        display_menu(menu,"a")
        choice = input("-> ")
        if choice not in data:
            continue
        else:
            return choice


def sauvegarder_commande(file_name: str, data: list):
    with open(file_name, 'a') as file:
        for item in data:
            str_item = [str(i) for i in item]
            file.write(','.join(str_item) + '\n')

def generer_numero_vente():
    return datetime.now().strftime("%d%m%Y%H%M%S")

def enregistrer_vente(vente_file: str, numero_vente: str, data: list):
    with open(vente_file, 'a') as file:
        for item in data:
            str_item = [str(i) for i in item]
            file.write(f"{numero_vente},{','.join(str_item)}\n")

def generer_facture(numero_commande: str, panier: list):
    with open(f"Fact_{numero_commande}.txt", "w") as facture:
        # name = "**********************FACTURE***********************\n"
        # facture.write(name)
        header = "Id,Libelle,Qte,PVu,Total\n"
        facture.write(header)
        somme = 0
        for item in panier:
            item.pop(4)
            somme += item[4]
            str_item = [str(i) for i in item]
            somme_str = ["-", "-", "-", "Total", str(somme)]
            facture.write(f"{','.join(str_item)}\n")
        facture.write(f"{','.join(somme_str)}")
    


def valider_panier(file_name_command: str, panier: list, all_products: list, file_name: str, shopping_cart_file: str, vente_file: str):
    numero_vente = generer_numero_vente()
    sauvegarder_commande(shopping_cart_file, panier)
    sauvegarder_commande(file_name_command, panier)
    for cmd in panier:
        id_produit = cmd[0]
        qte_commandee = int(cmd[2])
        update_product_database(id_produit, qte_commandee, all_products, file_name)
    enregistrer_vente(vente_file, numero_vente, panier)
    generer_facture(numero_vente, panier)



def traitement(panier: list, all_products: list, available_products: list) -> None:
    while True:
        display_product(panier)
        
        choice = input_choice('abc')
        if choice == 'a':
            valider_panier(COMMAND_LINE, panier, all_products, PRODUCT_FILE, PANIER, SALE_FILE)
            print("Commande validée avec succès!!!")
            FACTURE_PATH = f'fact_{generer_numero_vente()}.txt'
            FACTURE_PATH_2 = f'fact_{generer_numero_vente()}.xlsx'
            facture = get_file_content(FACTURE_PATH)
            display_product(facture)

            # le chemin actuel du fichier
            source = f'fact_{generer_numero_vente()}.txt'
            source2 = FACTURE_PATH_2

            # le chemin du dossier de destination
            destination = 'Factures/'
            destination2 = 'FacturesXLSX/'
            destination3 = 'FacturesPDF/'

            # Déplacement du fichier vers le dossier de destination
            shutil.move(source, destination)
            create_Facture_xlsx(facture,FACTURE_PATH_2, f"Votre Fact_{generer_numero_vente()}", generer_numero_vente() )
            shutil.move(source2, destination2)
            FACTURE_PATH_3 = f'fact_{generer_numero_vente()}.pdf'
            shutil.move(FACTURE_PATH_3, destination3)
            break

        elif choice == 'b':
            annuler_commande(panier)
            break
        else:
            display_menu(menu2, 1)
            answer = get_answer_2(1, 3)
            modifier_commande(panier, available_products, answer, 0)
            continue
        # print("Commande validée avec succès!!!")